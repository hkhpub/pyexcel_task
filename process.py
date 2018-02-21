import openpyxl

wb = openpyxl.load_workbook(filename="qa_category.xlsx")
# sheet_ranges = wb.sheetnames
print(wb.sheetnames)
hp_sheet = wb.worksheets[0]
category_sheet = wb.worksheets[1]

print(hp_sheet.columns)
num_rows = len([1 for row in hp_sheet.rows])
num_cols = len([1 for col in hp_sheet.columns])

data_entries = []
for row in range(num_rows):
    entry = dict()
    cell_values = []
    for col in range(num_cols):
        cell_values += [hp_sheet.cell(column=col+1, row=row+1).value]

    if cell_values[2] is None and cell_values[3] is not None:      # Question is None
        print(row, cell_values[3])

    # if row == 500:
    #     break
